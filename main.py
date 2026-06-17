import os, io, base64, json, shutil, uuid, re, zipfile
import pdfplumber
from pathlib import Path
from datetime import datetime
from typing import Optional

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pypdf import PdfReader
from groq import Groq
from PIL import Image
import aiofiles

import auth
import prompt as P
from admin import router as admin_router
from dotenv import load_dotenv

load_dotenv()
auth.init_db()

app = FastAPI(title="QC Compliance Checker")
app.include_router(admin_router)
app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
templates = Jinja2Templates(directory="templates")

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
SESSION_COOKIE = "qc_token"


def get_current_user(request: Request):
    token = (
        request.headers.get("X-Auth-Token") or
        request.cookies.get(SESSION_COOKIE) or ""
    )
    return auth.get_user_by_token(token)


def get_qc_items() -> list[dict]:
    """Load active QC items from DB (source of truth)."""
    return auth.get_all_qc_items(active_only=True)

# --- OCR via Groq ---
def ocr_file_with_groq(filepath: Path, api_key: str, prompt: str = None) -> str:
    client = Groq(api_key=api_key)
    suffix = filepath.suffix.lower()

    if suffix == ".pdf":
        try:
            reader = PdfReader(str(filepath))
            text = "\n".join(p.extract_text() or "" for p in reader.pages)
            if text.strip():
                return text
        except:
            pass
        # fallback: rasterize first page and use vision
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(str(filepath), dpi=150, first_page=1, last_page=1)
            if images:
                buf = io.BytesIO()
                images[0].save(buf, format="JPEG")
                img_b64 = base64.b64encode(buf.getvalue()).decode()
                resp = client.chat.completions.create(
                    model="meta-llama/llama-4-scout-17b-16e-instruct",
                    messages=[{"role": "user", "content": [
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                        {"type": "text", "text": prompt or "Extract all text from this document image. Return plain text only."}
                    ]}],
                    max_tokens=2000
                )
                return resp.choices[0].message.content or ""
        except Exception as e:
            return f"[OCR error: {e}]"

    elif suffix in (".jpg", ".jpeg", ".png", ".webp"):
        img = Image.open(str(filepath))
        if img.mode != "RGB":
            img = img.convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        img_b64 = base64.b64encode(buf.getvalue()).decode()
        try:
            resp = client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[{"role": "user", "content": [
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                    {"type": "text", "text": prompt or "Extract all text from this document image. Return plain text only."}
                ]}],
                max_tokens=2000
            )
            return resp.choices[0].message.content or ""
        except Exception as e:
            return f"[OCR error: {e}]"

    return ""


def check_document_against_agreement(doc_text: str, agreement_text: str, check_key: str, client: Groq) -> dict:
    """Use Groq LLM to check if document satisfies a specific QC item."""
    item = auth.get_qc_item_by_key(check_key)
    label = item["label"] if item else check_key

    prompt = P.GENERIC_QC_PROMPT.format(
        agreement_text=agreement_text[:3000],
        doc_text=doc_text[:2000],
        label=label,
    )

    resp = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=200,
        temperature=0.1
    )
    raw = resp.choices[0].message.content or ""
    raw = raw.replace("```json", "").replace("```", "").strip()
    try:
        return json.loads(raw)
    except:
        return {"result": "N/A", "remark": raw[:120]}


def extract_agreement_text(filepath: Path) -> str:
    # Try pdfplumber first (better text extraction), fall back to pypdf
    try:
        pages = []
        with pdfplumber.open(str(filepath)) as pdf:
            for page in pdf.pages:
                pages.append(page.extract_text() or "")
        text = "\n\n".join(pages)
        if text.strip():
            return text
    except Exception:
        pass
    try:
        reader = PdfReader(str(filepath))
        return "\n".join(p.extract_text() or "" for p in reader.pages)
    except Exception:
        return ""


def _clean_address_cell(raw: str, customer_name: str) -> str:
    """Strip customer name from top of a pdfplumber table cell and normalise."""
    lines = [l.strip() for l in raw.splitlines() if l.strip()]
    if lines and customer_name:
        all_names = [customer_name] + [n.strip() for n in re.split(r'\s*&\s*|\s+and\s+', customer_name, flags=re.IGNORECASE)]
        if any(lines[0].lower() == n.lower() for n in all_names):
            lines = lines[1:]
    addr = ", ".join(lines)
    addr = re.sub(r',\s*,', ',', addr)
    return addr.strip(", ")


def _parse_addresses_from_pdf(pdf_path: Path) -> dict:
    """
    Extract customer name and addresses directly from pdfplumber table cells.
    This avoids all text-merging hallucination issues — each cell is cleanly separated.
    Table 1 on page 2 of ADS Solar agreements has rows:
      ['Customer name:', 'Christine Hill & Ashley Hill', ...]
      ['Billing address:', 'Christine Hill & Ashley Hill\\n5 Elm Cres,\\nEmerald VIC 3782\\nAustralia', ...]
    """
    result = {}
    try:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables():
                    for row in table:
                        if not row or len(row) < 2:
                            continue
                        label = (row[0] or "").strip().lower().rstrip(":")
                        value = (row[1] or "").strip()
                        if not value:
                            continue
                        if label == "customer name":
                            result["customer_name"] = value
                        elif label == "billing address":
                            result["billing_address"] = _clean_address_cell(value, result.get("customer_name", ""))
                        elif label == "delivery address":
                            result["delivery_address"] = _clean_address_cell(value, result.get("customer_name", ""))
                if result.get("billing_address"):
                    break
    except Exception:
        pass

    if result.get("billing_address"):
        result["installation_address"] = result["billing_address"]
    return result


def extract_quote_with_groq(text: str, api_key: str) -> dict:
    """Use Groq LLM to extract all structured fields from agreement text."""
    client = Groq(api_key=api_key)
    resp = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        temperature=0,
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "system",
                "content": P.QUOTE_EXTRACTION_SYSTEM + "\n\n" + P.QUOTE_SCHEMA,
            },
            {"role": "user", "content": "Document text:\n\n" + text[:6000]},
        ],
    )
    return json.loads(resp.choices[0].message.content)


def generate_excel(session_dir: Path, results: list, meta: dict) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QC Checklist"

    orange = "E8520A"
    dark = "1A1A2E"
    light_gray = "F5F5F5"
    yes_green = "D4EDDA"
    no_red = "F8D7DA"
    na_yellow = "FFF3CD"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 50

    # Header block
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "JOB COMPLIANCE CHECKLIST (QC)"
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=orange)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    fields = [
        ("Customer Name", meta.get("customer_name", "")),
        ("Address", meta.get("address", "")),
        ("Checked By", meta.get("checked_by", "")),
        ("Date", meta.get("date", "")),
        ("Job Details", meta.get("job_details", "")),
    ]
    for i, (k, v) in enumerate(fields, 2):
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:E{i}")
        ws[f"A{i}"].value = k + " :"
        ws[f"A{i}"].font = Font(name="Arial", bold=True, size=10)
        ws[f"A{i}"].alignment = Alignment(vertical="center")
        ws[f"C{i}"].value = v
        ws[f"C{i}"].font = Font(name="Arial", size=10)
        ws.row_dimensions[i].height = 18

    header_row = len(fields) + 2
    for col, hdr in zip(["A","B","C","D","E"], ["Sno.", "CheckList", "Yes/No", "File", "Remarks"]):
        c = ws[f"{col}{header_row}"]
        c.value = hdr
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=dark)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[header_row].height = 22

    for idx, r in enumerate(results, header_row + 1):
        result_val = r.get("result", "N/A")
        row_fill = PatternFill("solid", fgColor=(
            yes_green if result_val == "Yes" else
            no_red if result_val == "No" else
            na_yellow
        ))
        data = [r.get("sno",""), r.get("label",""), result_val, r.get("filename",""), r.get("remark","")]
        for col, val in zip(["A","B","C","D","E"], data):
            c = ws[f"{col}{idx}"]
            c.value = val
            c.font = Font(name="Arial", size=9)
            c.fill = row_fill
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=(col == "E"))
        ws.row_dimensions[idx].height = 16

    out_path = session_dir / "QC_Results.xlsx"
    wb.save(str(out_path))
    return out_path


# ── Routes ──────────────────────────────────────────────────────────────────

@app.get("/api/qc-items")
async def api_qc_items():
    """Return active QC checklist items from the database."""
    return JSONResponse(get_qc_items())


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    qc_items = get_qc_items()
    return templates.TemplateResponse(request=request, name="index.html", context={"qc_items": qc_items, "user": {"name": ""}})


@app.post("/api/upload-agreement")
async def upload_agreement(request: Request, file: UploadFile = File(...)):
    user = get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    session_id = str(uuid.uuid4())[:8]
    sess_dir = UPLOAD_DIR / session_id
    sess_dir.mkdir(parents=True, exist_ok=True)
    dest = sess_dir / "agreement.pdf"
    pdf_bytes = await file.read()
    async with aiofiles.open(dest, "wb") as f:
        await f.write(pdf_bytes)
    text = extract_agreement_text(dest)

    # Use Groq to extract all structured fields from the agreement
    api_key = GROQ_API_KEY
    quote = {}
    if api_key:
        try:
            quote = extract_quote_with_groq(text, api_key)
        except Exception:
            quote = {}

    # Override address/name fields using pdfplumber table extraction — no hallucination
    parsed_addresses = _parse_addresses_from_pdf(dest)
    for field, value in parsed_addresses.items():
        if value:
            quote[field] = value

    # Pull the key fields we need for the project record
    customer_name = quote.get("customer_name") or ""
    address       = quote.get("billing_address") or quote.get("delivery_address") or ""
    job_details   = ""
    # Build job_details from line items
    items = quote.get("line_items") or []
    if items:
        job_details = " | ".join(
            f"{i.get('quantity', '')}x {i.get('specification', '')}"
            for i in items[:3]
        )

    auth.upsert_project(session_id, user["id"], customer_name, address, file.filename)
    auth.save_agreement_pdf(session_id, pdf_bytes)
    # Save full extracted quote as JSON for later use
    if quote:
        auth.save_quote_data(session_id, json.dumps(quote))

    return JSONResponse({
        "session_id":   session_id,
        "filename":     file.filename,
        "customer_name": customer_name,
        "address":      address,
        "job_details":  job_details,
        # Rich fields for UI display
        "quote_number":        quote.get("quote_number") or "",
        "customer_phone":      quote.get("customer_phone") or "",
        "customer_email":      quote.get("customer_email") or "",
        "total_price":         quote.get("total_price"),
        "deposit":             quote.get("deposit"),
        "system_price":        quote.get("system_price"),
        "proposed_install_date": quote.get("proposed_install_date") or "",
        "roof_type":           quote.get("roof_type") or "",
        "inverter_phase":      quote.get("inverter_phase") or "",
        "stories":             quote.get("stories") or "",
        "signed_by":           quote.get("signed_by") or "",
        "signed_date":         quote.get("signed_date") or "",
        "line_items":          items,
        "retailer_name":       quote.get("retailer_name") or "",
    })


# Rules live in prompt.py — edit that file to add/change per-item logic.


@app.post("/api/check-document")
async def check_document(
    session_id: str = Form(...),
    check_key: str = Form(...),
    checked_by: str = Form(""),
    groq_api_key: str = Form(""),
    file: UploadFile = File(...)
):
    sess_dir = UPLOAD_DIR / session_id
    sess_dir.mkdir(parents=True, exist_ok=True)

    agreement_path = sess_dir / "agreement.pdf"
    if not agreement_path.exists():
        # Try to restore from DB blob (portability: uploads folder may be missing)
        pdf_blob = auth.get_agreement_pdf(session_id)
        if not pdf_blob:
            raise HTTPException(404, "Agreement PDF not found. Please upload the agreement again.")
        agreement_path.write_bytes(pdf_blob)

    suffix = Path(file.filename).suffix.lower()
    safe_name = f"{check_key}{suffix}"
    dest = sess_dir / safe_name
    async with aiofiles.open(dest, "wb") as f:
        await f.write(await file.read())

    item = auth.get_qc_item_by_key(check_key) or {}
    agreement_text = extract_agreement_text(agreement_path)

    project    = auth.get_project(session_id) or {}
    quote_data = json.loads(project["quote_data"]) if project.get("quote_data") else {}

    api_key = groq_api_key or GROQ_API_KEY

    def _save_and_return(result_val: str, remark: str, ocr_preview: str = "") -> JSONResponse:
        auth.upsert_check_result(session_id, check_key, result_val, remark, file.filename, ocr_preview)
        return JSONResponse({
            "sno":         item.get("sno", ""),
            "label":       item.get("label", check_key),
            "key":         check_key,
            "filename":    file.filename,
            "result":      result_val,
            "remark":      remark,
            "ocr_preview": ocr_preview,
        })

    if check_key in P.NO_OCR_KEYS:
        custom = P.run_item_rule(check_key, file.filename, "", agreement_text, quote_data)
        return _save_and_return(custom["result"], custom["remark"])

    if check_key in P.OCR_RULE_KEYS:
        if not api_key:
            raise HTTPException(400, "Groq API key required for OCR")
        if check_key in ("storey_roof", "scissor_lift", "tilt_frame"):
            doc_text = ocr_file_with_groq(dest, api_key, prompt=P.STOREY_PROMPT)
        else:
            doc_text = ocr_file_with_groq(dest, api_key)
        custom = P.run_item_rule(check_key, file.filename, doc_text, agreement_text, quote_data)
        return _save_and_return(custom["result"], custom["remark"], doc_text[:200])

    if not api_key:
        raise HTTPException(400, "Groq API key required for this check")

    client   = Groq(api_key=api_key)
    doc_text = ocr_file_with_groq(dest, api_key)
    result   = check_document_against_agreement(doc_text, agreement_text, check_key, client)
    return _save_and_return(result.get("result", "N/A"), result.get("remark", ""), doc_text[:200])


@app.post("/api/generate-excel")
async def generate_excel_report(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    body = await request.json()
    session_id = body.get("session_id")
    meta       = body.get("meta", {})

    sess_dir = UPLOAD_DIR / session_id
    sess_dir.mkdir(exist_ok=True)

    # Build results from DB — merge with active QC items list so every item appears
    qc_items    = get_qc_items()
    db_results  = {r["item_key"]: r for r in auth.get_check_results(session_id)}
    results = []
    for item in qc_items:
        row = db_results.get(item["key"])
        results.append({
            "sno":      item["sno"],
            "label":    item["label"],
            "key":      item["key"],
            "result":   row["result"]   if row else "—",
            "remark":   row["remark"]   if row else "",
            "filename": row["filename"] if row else "",
        })

    out_path = generate_excel(sess_dir, results, meta)
    return JSONResponse({"download_url": f"/uploads/{session_id}/QC_Results.xlsx"})


@app.get("/download/{session_id}")
async def download_excel(session_id: str):
    path = UPLOAD_DIR / session_id / "QC_Results.xlsx"
    if not path.exists():
        raise HTTPException(404, "File not found")
    return FileResponse(str(path), filename="QC_Results.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/upload-excel")
async def upload_excel(
    request: Request,
    file: UploadFile = File(...),
    session_id: str = Form(""),
):
    """
    Import a pre-filled QC checklist Excel file.
    If session_id is provided, attaches results to that existing session
    (used after the PDF agreement is already uploaded).
    Otherwise creates a new session.
    """
    user = get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".xlsx", ".xls"):
        raise HTTPException(400, "Only .xlsx / .xls files are supported")

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel file: {e}")

    ws = wb.active

    qc_items  = get_qc_items()
    label_map = {item["label"].strip().lower(): item for item in qc_items}
    key_map   = {item["key"].strip().lower():   item for item in qc_items}

    if not session_id:
        session_id = str(uuid.uuid4())[:8]

    sess_dir = UPLOAD_DIR / session_id
    sess_dir.mkdir(parents=True, exist_ok=True)

    customer_name = ""
    address       = ""
    imported      = []

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        cells = [str(c).strip() if c is not None else "" for c in row]

        # Pick up customer/address from header block rows
        if len(cells) >= 2:
            label_cell = cells[0].lower().rstrip(" :")
            if label_cell in ("customer name", "customer") and cells[1]:
                customer_name = cells[1]
            elif label_cell == "address" and cells[1]:
                address = cells[1]

        # Detect data rows by matching key or label in first two cols
        item = None
        for col_idx in (0, 1):
            if col_idx >= len(cells):
                continue
            val = cells[col_idx].strip().lower()
            if val in key_map:
                item = key_map[val]
                break
            if val in label_map:
                item = label_map[val]
                break

        if not item:
            continue

        result_val = ""
        remark     = ""
        filename   = ""
        for idx in (2, 3):
            if idx < len(cells) and cells[idx].lower() in ("yes", "no", "n/a", "—", "-"):
                result_val = cells[idx].strip()
                if result_val in ("—", "-"):
                    result_val = "N/A"
                remark   = cells[idx + 1] if idx + 1 < len(cells) else ""
                filename = cells[idx + 2] if idx + 2 < len(cells) else ""
                break

        if not result_val:
            continue

        auth.upsert_check_result(session_id, item["key"], result_val, remark, filename, "")
        imported.append({
            "sno":      item["sno"],
            "label":    item["label"],
            "key":      item["key"],
            "result":   result_val,
            "remark":   remark,
            "filename": filename,
        })

    # Only upsert project if no session existed (standalone Excel-only flow)
    if not auth.get_project(session_id):
        auth.upsert_project(session_id, user["id"], customer_name, address, file.filename)

    return JSONResponse({
        "ok":           True,
        "session_id":   session_id,
        "customer_name": customer_name,
        "address":      address,
        "results":      imported,
        "imported":     len(imported),
    })


@app.post("/api/mark-na")
async def mark_na(request: Request):
    """Persist a manual N/A mark for a checklist item."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    body       = await request.json()
    session_id = body.get("session_id", "")
    item_key   = body.get("item_key", "")
    remark     = body.get("remark", "Marked N/A manually")
    if not session_id or not item_key:
        raise HTTPException(400, "session_id and item_key required")
    auth.upsert_check_result(session_id, item_key, "N/A", remark, "", "")
    return JSONResponse({"ok": True})


@app.post("/api/sync-project")
async def sync_project(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    body       = await request.json()
    session_id = body.get("session_id", "")
    auth.upsert_project(
        session_id, user["id"],
        body.get("customer", ""), body.get("address", ""), body.get("agreement", ""),
    )
    return JSONResponse({"ok": True})


@app.get("/api/my-last-project")
async def my_last_project(request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")
    projects = auth.get_projects_by_user(user["id"])
    if not projects:
        return JSONResponse({"ok": False})
    p = projects[0]  # most recent

    # Compute counts from check_results
    check_rows = auth.get_check_results(p["session_id"])
    yes   = sum(1 for r in check_rows if r["result"] == "Yes")
    no    = sum(1 for r in check_rows if r["result"] == "No")
    na    = sum(1 for r in check_rows if r["result"] == "N/A")
    total = len(check_rows)

    # Return per-item results so the frontend can fully restore the checklist
    results_map = {r["item_key"]: {
        "key":      r["item_key"],
        "result":   r["result"],
        "remark":   r["remark"],
        "filename": r["filename"],
        "ocr_preview": r.get("ocr_preview", ""),
    } for r in check_rows}

    # Merge with current QC items to fill in sno/label
    qc_items = get_qc_items()
    for item in qc_items:
        if item["key"] in results_map:
            results_map[item["key"]]["sno"]   = item["sno"]
            results_map[item["key"]]["label"]  = item["label"]

    return JSONResponse({
        "ok":         True,
        "session_id": p["session_id"],
        "customer":   p["customer"]  or "",
        "address":    p["address"]   or "",
        "agreement":  p["agreement"] or "",
        "date":       (p["created_at"] or "")[:10],
        "yes":        yes,
        "no":         no,
        "na":         na,
        "total":      total,
        "results":    list(results_map.values()),
    })


@app.delete("/api/session/{session_id}")
async def delete_session_dir(session_id: str):
    sess_dir = UPLOAD_DIR / session_id
    if sess_dir.exists():
        shutil.rmtree(sess_dir)
    auth.delete_check_results(session_id)
    return JSONResponse({"status": "deleted", "session_id": session_id})


@app.post("/api/batch-check")
async def batch_check(
    request: Request,
    session_id: str = Form(...),
    zip_file: UploadFile = File(...),
):
    """
    Accept a ZIP of supporting documents, map each file to a QC check key by
    filename stem, run the same logic as /api/check-document for each, and
    return a list of results.
    """
    user = get_current_user(request)
    if not user:
        raise HTTPException(401, "Not authenticated")

    sess_dir = UPLOAD_DIR / session_id
    sess_dir.mkdir(parents=True, exist_ok=True)

    agreement_path = sess_dir / "agreement.pdf"
    if not agreement_path.exists():
        pdf_blob = auth.get_agreement_pdf(session_id)
        if not pdf_blob:
            raise HTTPException(404, "Agreement PDF not found. Upload the agreement first.")
        agreement_path.write_bytes(pdf_blob)

    zip_bytes = await zip_file.read()
    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        raise HTTPException(400, "Uploaded file is not a valid ZIP archive.")

    agreement_text = extract_agreement_text(agreement_path)
    api_key = GROQ_API_KEY

    project    = auth.get_project(session_id) or {}
    quote_data = json.loads(project["quote_data"]) if project.get("quote_data") else {}

    # Build a key→item map from DB (dynamic — reflects any admin edits)
    valid_keys = {item["key"]: item for item in get_qc_items()}

    # Files that should run two checks from one image
    DUAL_KEY_MAP = {
        "roof_pic": ["roof_pic", "storey_roof", "scissor_lift", "tilt_frame"],
    }

    def _resolve_keys(stem_lower: str) -> list[str]:
        """Return all QC keys for a filename stem. Most files map to one key;
        house/roof images map to both roof_pic and storey_roof."""
        if stem_lower in valid_keys:
            matched = stem_lower
        else:
            matched = None
            for keywords, key in P.KEYWORD_MAP:
                for kw in keywords:
                    if kw in stem_lower:
                        matched = key if key in valid_keys else None
                        break
                if matched:
                    break
        if not matched:
            return []
        return [k for k in DUAL_KEY_MAP.get(matched, [matched]) if k in valid_keys]

    results_list = []
    skipped = []

    for name in zf.namelist():
        if name.endswith("/") or name.startswith("__MACOSX") or "/." in name:
            continue

        stem   = Path(name).stem.lower()
        suffix = Path(name).suffix.lower()

        if suffix not in P.VALID_EXTENSIONS:
            skipped.append({"filename": name, "reason": "unsupported extension"})
            continue

        check_keys = _resolve_keys(stem)
        if not check_keys:
            skipped.append({"filename": name, "reason": "filename does not match any checklist key"})
            continue

        original_filename = Path(name).name
        file_bytes = zf.read(name)

        # OCR the image once if any of the keys need it, reuse for all keys
        cached_ocr: dict[str, str] = {}

        for check_key in check_keys:
            item = valid_keys[check_key]

            dest = sess_dir / f"{check_key}{suffix}"
            dest.write_bytes(file_bytes)

            def _save(result_val: str, remark: str, ocr_preview: str = "",
                      _key=check_key, _item=item) -> dict:
                auth.upsert_check_result(session_id, _key, result_val, remark, original_filename, ocr_preview)
                return {
                    "sno":         _item["sno"],
                    "label":       _item["label"],
                    "key":         _key,
                    "filename":    original_filename,
                    "result":      result_val,
                    "remark":      remark,
                    "ocr_preview": ocr_preview,
                }

            try:
                if check_key in P.NO_OCR_KEYS:
                    custom = P.run_item_rule(check_key, original_filename, "", agreement_text, quote_data)
                    results_list.append(_save(custom["result"], custom["remark"]))

                elif check_key in P.OCR_RULE_KEYS:
                    if not api_key:
                        results_list.append(_save("N/A", "Groq API key not configured — skipped"))
                        continue
                    ocr_prompt = P.STOREY_PROMPT if check_key in ("storey_roof", "scissor_lift", "tilt_frame") else None
                    cache_key  = ocr_prompt or "__default__"
                    if cache_key not in cached_ocr:
                        cached_ocr[cache_key] = ocr_file_with_groq(dest, api_key, prompt=ocr_prompt)
                    doc_text = cached_ocr[cache_key]
                    custom = P.run_item_rule(check_key, original_filename, doc_text, agreement_text, quote_data)
                    results_list.append(_save(custom["result"], custom["remark"], doc_text[:200]))

                else:
                    if not api_key:
                        results_list.append(_save("N/A", "Groq API key not configured — skipped"))
                        continue
                    client = Groq(api_key=api_key)
                    if "__default__" not in cached_ocr:
                        cached_ocr["__default__"] = ocr_file_with_groq(dest, api_key)
                    doc_text = cached_ocr["__default__"]
                    result   = check_document_against_agreement(doc_text, agreement_text, check_key, client)
                    results_list.append(_save(result.get("result", "N/A"), result.get("remark", ""), doc_text[:200]))

            except Exception as e:
                results_list.append(_save("N/A", f"Error during check: {str(e)[:120]}"))

    zf.close()
    return JSONResponse({"results": results_list, "skipped": skipped})